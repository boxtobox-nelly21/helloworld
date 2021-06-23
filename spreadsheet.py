import openpyxl

# This is supposed to read a spreadsheet from the local directory and then automate its processing.
# This is useful when you have lots of data so that your brain doesnt go Brrrr!!

# openpyxl package allows you to work with spreadsheets specifically instead of io module thats in built

# var initializations which are to be populated later on
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]
products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

# Sequentially goes through each row checking the supplier name and grouping them. At this point am just flexing.
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # calculating the number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculating the total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

inv_file.save("inventory_with_total_value.xlsx")
